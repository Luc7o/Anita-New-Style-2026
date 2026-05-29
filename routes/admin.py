import os, uuid, json, re
from functools import wraps
from datetime import datetime, timedelta
from flask import (Blueprint, render_template, redirect, url_for, flash,
                   request, current_app, jsonify, Response)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from sqlalchemy import func
from app_extensions import db
from models.producto  import Producto, Categoria, ImagenProducto, Proveedor
from models.pedido    import Pedido, DetallePedido
from models.usuario   import Usuario
from models.movimiento import MovimientoStock
from models.venta     import VentaFisica, DetalleVenta
from forms.almacen_forms import (FormProductoAlmacen, FormCategoria,
                                  FormProveedor, FormMovimiento, FormVenta)

bp = Blueprint('admin', __name__)


# ═══════════════════════════════════════════════════════════════════════════════
# DECORADOR
# ═══════════════════════════════════════════════════════════════════════════════
def admin_requerido(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.es_admin:
            flash('Acceso denegado.', 'danger')
            return redirect(url_for('tienda.inicio'))
        return f(*args, **kwargs)
    return decorated


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def _guardar_imagen(archivo):
    ext = archivo.filename.rsplit('.', 1)[-1].lower() if '.' in archivo.filename else ''
    if ext not in current_app.config.get('ALLOWED_EXTENSIONS', {'png', 'jpg', 'jpeg', 'webp'}):
        raise ValueError(f'Tipo de archivo no permitido: .{ext}')
    nombre = f"prod_{uuid.uuid4().hex[:12]}.{ext}"
    ruta   = os.path.join(current_app.config['UPLOAD_FOLDER'], nombre)
    archivo.save(ruta)
    return nombre


def _slugify(texto):
    reemplazos = {
        'á':'a','é':'e','í':'i','ó':'o','ú':'u','ñ':'n',
        'ä':'a','ë':'e','ï':'i','ö':'o','ü':'u',
        ' ':'-','&':'-','/':'-','_':'-',
    }
    t = texto.lower()
    for k, v in reemplazos.items():
        t = t.replace(k, v)
    slug = ''.join(c for c in t if c.isalnum() or c == '-')
    slug = re.sub(r'-{2,}', '-', slug).strip('-')
    return slug


# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/')
@admin_requerido
def dashboard():
    hoy   = datetime.utcnow().date()
    mes_i = datetime(hoy.year, hoy.month, 1)

    # ── KPIs generales ──────────────────────────────────────────────────────
    total_productos  = Producto.query.filter_by(activo=True).count()
    total_clientes   = Usuario.query.filter_by(es_admin=False).count()
    sin_stock        = Producto.query.filter(Producto.stock <= 0, Producto.activo==True).count()
    productos_stock_bajo = Producto.query.filter(
        Producto.stock > 0, Producto.stock <= Producto.stock_minimo, Producto.activo==True
    ).count()

    # ── Ventas físicas ───────────────────────────────────────────────────────
    ventas_hoy = db.session.query(func.sum(VentaFisica.total))\
        .filter(func.date(VentaFisica.fecha) == hoy, VentaFisica.anulada==False).scalar() or 0
    ventas_mes = db.session.query(func.sum(VentaFisica.total))\
        .filter(VentaFisica.fecha >= mes_i, VentaFisica.anulada==False).scalar() or 0
    num_ventas_hoy = VentaFisica.query\
        .filter(func.date(VentaFisica.fecha) == hoy, VentaFisica.anulada==False).count()

    # ── Ventas online (pedidos) ──────────────────────────────────────────────
    total_ventas_online = db.session.query(func.sum(Pedido.total))\
        .filter(Pedido.estado_pago=='pagado').scalar() or 0
    pedidos_pendientes  = Pedido.query.filter_by(estado='pendiente').count()
    ultimos_pedidos     = Pedido.query.order_by(Pedido.fecha_creacion.desc()).limit(6).all()

    # ── Inventario ───────────────────────────────────────────────────────────
    valor_inventario = db.session.query(
        func.sum(Producto.precio * Producto.stock)
    ).filter(Producto.activo==True).scalar() or 0

    # ── Alertas de stock ─────────────────────────────────────────────────────
    alertas_stock = Producto.query.filter(
        Producto.stock <= Producto.stock_minimo, Producto.activo==True
    ).order_by(Producto.stock.asc()).limit(8).all()

    # ── Últimas ventas físicas ───────────────────────────────────────────────
    ultimas_ventas = VentaFisica.query\
        .filter_by(anulada=False)\
        .order_by(VentaFisica.fecha.desc()).limit(6).all()

    # ── Últimos movimientos ──────────────────────────────────────────────────
    ultimos_movimientos = MovimientoStock.query\
        .order_by(MovimientoStock.fecha.desc()).limit(6).all()

    # ── Ventas últimos 7 días ────────────────────────────────────────────────
    ventas_semana = []
    for i in range(6, -1, -1):
        dia = hoy - timedelta(days=i)
        total = db.session.query(func.sum(VentaFisica.total))\
            .filter(func.date(VentaFisica.fecha) == dia, VentaFisica.anulada==False)\
            .scalar() or 0
        ventas_semana.append({'dia': dia.strftime('%d/%m'), 'total': float(total)})

    # ── Top productos del mes ────────────────────────────────────────────────
    top_productos = db.session.query(
        Producto.nombre,
        func.sum(DetalleVenta.cantidad).label('unidades'),
        func.sum(DetalleVenta.subtotal).label('monto')
    ).join(DetalleVenta, DetalleVenta.producto_id == Producto.id)\
     .join(VentaFisica, VentaFisica.id == DetalleVenta.venta_id)\
     .filter(VentaFisica.fecha >= mes_i, VentaFisica.anulada==False)\
     .group_by(Producto.id, Producto.nombre)\
     .order_by(func.sum(DetalleVenta.cantidad).desc())\
     .limit(5).all()

    # ── Ventas por categoría (online) ────────────────────────────────────────
    ventas_cat = db.session.query(
        Categoria.nombre,
        func.sum(DetallePedido.subtotal).label('total')
    ).join(Producto, Producto.id==DetallePedido.producto_id)\
     .join(Categoria, Categoria.id==Producto.categoria_id)\
     .join(Pedido, Pedido.id==DetallePedido.pedido_id)\
     .filter(Pedido.estado_pago=='pagado')\
     .group_by(Categoria.nombre).all()

    return render_template('admin/dashboard.html',
        total_productos=total_productos, total_clientes=total_clientes,
        sin_stock=sin_stock, productos_stock_bajo=productos_stock_bajo,
        ventas_hoy=ventas_hoy, ventas_mes=ventas_mes, num_ventas_hoy=num_ventas_hoy,
        total_ventas_online=total_ventas_online, pedidos_pendientes=pedidos_pendientes,
        valor_inventario=valor_inventario,
        alertas_stock=alertas_stock, ultimas_ventas=ultimas_ventas,
        ultimos_movimientos=ultimos_movimientos, ultimos_pedidos=ultimos_pedidos,
        ventas_semana=ventas_semana, top_productos=top_productos,
        ventas_cat=ventas_cat,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# PRODUCTOS
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/productos')
@admin_requerido
def productos():
    pagina   = request.args.get('pagina', 1, type=int)
    busqueda = request.args.get('q', '')
    cat_id   = request.args.get('cat', 0, type=int)
    estado   = request.args.get('estado', 'activos')

    query = Producto.query
    if estado == 'activos':
        query = query.filter_by(activo=True)
    elif estado == 'inactivos':
        query = query.filter_by(activo=False)
    if busqueda:
        query = query.filter(
            (Producto.nombre.ilike(f'%{busqueda}%')) |
            (Producto.sku.ilike(f'%{busqueda}%')) |
            (Producto.codigo_barras.ilike(f'%{busqueda}%'))
        )
    if cat_id:
        query = query.filter_by(categoria_id=cat_id)

    prods      = query.order_by(Producto.nombre.asc()).paginate(page=pagina, per_page=20)
    categorias = Categoria.query.filter_by(activo=True).all()
    return render_template('admin/productos.html', productos=prods,
                           busqueda=busqueda, categorias=categorias,
                           cat_sel=cat_id, estado=estado)


@bp.route('/productos/nuevo', methods=['GET', 'POST'])
@admin_requerido
def nuevo_producto():
    form = FormProductoAlmacen()
    form.categoria_id.choices = [(c.id, c.nombre) for c in Categoria.query.filter_by(activo=True).all()]
    form.proveedor_id.choices = [(0, '— Sin proveedor —')] + [
        (p.id, p.nombre) for p in Proveedor.query.filter_by(activo=True).all()]
    if form.validate_on_submit():
        tallas  = json.dumps([t.strip() for t in form.tallas.data.split(',') if t.strip()]) if form.tallas.data else None
        colores = json.dumps([c.strip() for c in form.colores.data.split(',') if c.strip()]) if form.colores.data else None
        prod = Producto(
            nombre=form.nombre.data, descripcion=form.descripcion.data,
            sku=form.sku.data or None,
            codigo_barras=form.codigo_barras.data or None,
            categoria_id=form.categoria_id.data,
            proveedor_id=form.proveedor_id.data if form.proveedor_id.data else None,
            precio_compra=form.precio_compra.data or 0,
            precio=form.precio.data,
            precio_oferta=form.precio_oferta.data or None,
            stock=form.stock.data or 0,
            stock_minimo=form.stock_minimo.data or 5,
            tallas=tallas, colores=colores,
            unidad=form.unidad.data,
            destacado=form.destacado.data,
            es_nuevo=form.es_nuevo.data,
            activo=form.activo.data,
        )
        if form.imagen.data and form.imagen.data.filename:
            try:
                prod.imagen_principal = _guardar_imagen(form.imagen.data)
            except ValueError as e:
                flash(str(e), 'danger')
                return render_template('admin/producto_form.html', form=form, titulo='Nuevo producto')
        db.session.add(prod)
        db.session.flush()
        if prod.stock > 0:
            mov = MovimientoStock(
                producto_id=prod.id, tipo='entrada',
                cantidad=prod.stock, stock_antes=0, stock_despues=prod.stock,
                motivo='Stock inicial al registrar producto',
                usuario_id=current_user.id
            )
            db.session.add(mov)
        db.session.commit()
        flash('Producto creado exitosamente.', 'success')
        return redirect(url_for('admin.productos'))
    return render_template('admin/producto_form.html', form=form, titulo='Nuevo producto')


@bp.route('/productos/<int:id>/editar', methods=['GET', 'POST'])
@admin_requerido
def editar_producto(id):
    prod = Producto.query.get_or_404(id)
    form = FormProductoAlmacen(obj=prod)
    form.categoria_id.choices = [(c.id, c.nombre) for c in Categoria.query.filter_by(activo=True).all()]
    form.proveedor_id.choices = [(0, '— Sin proveedor —')] + [
        (p.id, p.nombre) for p in Proveedor.query.filter_by(activo=True).all()]
    if request.method == 'GET':
        form.tallas.data  = ', '.join(prod.tallas_lista)
        form.colores.data = ', '.join(prod.colores_lista)
        form.proveedor_id.data = prod.proveedor_id or 0
    if form.validate_on_submit():
        prod.nombre        = form.nombre.data
        prod.descripcion   = form.descripcion.data
        prod.sku           = form.sku.data or None
        prod.codigo_barras = form.codigo_barras.data or None
        prod.categoria_id  = form.categoria_id.data
        prod.proveedor_id  = form.proveedor_id.data if form.proveedor_id.data else None
        prod.precio_compra = form.precio_compra.data or 0
        prod.precio        = form.precio.data
        prod.precio_oferta = form.precio_oferta.data or None
        prod.stock_minimo  = form.stock_minimo.data or 5
        prod.tallas  = json.dumps([t.strip() for t in form.tallas.data.split(',') if t.strip()]) if form.tallas.data else None
        prod.colores = json.dumps([c.strip() for c in form.colores.data.split(',') if c.strip()]) if form.colores.data else None
        prod.unidad  = form.unidad.data
        prod.destacado = form.destacado.data
        prod.es_nuevo  = form.es_nuevo.data
        prod.activo    = form.activo.data
        if form.imagen.data and form.imagen.data.filename:
            try:
                prod.imagen_principal = _guardar_imagen(form.imagen.data)
            except ValueError as e:
                flash(str(e), 'danger')
                return render_template('admin/producto_form.html', form=form, titulo='Editar producto', producto=prod)
        db.session.commit()
        flash('Producto actualizado correctamente.', 'success')
        return redirect(url_for('admin.productos'))
    return render_template('admin/producto_form.html', form=form, titulo='Editar producto', producto=prod)


@bp.route('/productos/<int:id>')
@admin_requerido
def ver_producto(id):
    prod        = Producto.query.get_or_404(id)
    movimientos = MovimientoStock.query.filter_by(producto_id=id)\
        .order_by(MovimientoStock.fecha.desc()).limit(20).all()
    return render_template('admin/producto_detalle.html', producto=prod, movimientos=movimientos)


@bp.route('/productos/<int:id>/toggle', methods=['POST'])
@admin_requerido
def toggle_producto(id):
    prod = Producto.query.get_or_404(id)
    prod.activo = not prod.activo
    db.session.commit()
    estado = 'activado' if prod.activo else 'desactivado'
    flash(f'Producto {estado}.', 'info')
    return redirect(url_for('admin.productos'))


@bp.route('/productos/<int:id>/eliminar', methods=['POST'])
@admin_requerido
def eliminar_producto(id):
    prod = Producto.query.get_or_404(id)
    prod.activo = False
    db.session.commit()
    flash('Producto desactivado.', 'warning')
    return redirect(url_for('admin.productos'))


# ═══════════════════════════════════════════════════════════════════════════════
# MOVIMIENTOS DE STOCK
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/movimientos')
@admin_requerido
def movimientos():
    pagina      = request.args.get('pagina', 1, type=int)
    tipo        = request.args.get('tipo', '')
    prod_id     = request.args.get('prod', 0, type=int)
    fecha_desde = request.args.get('desde', '')
    fecha_hasta = request.args.get('hasta', '')

    query = MovimientoStock.query
    if tipo:
        query = query.filter_by(tipo=tipo)
    if prod_id:
        query = query.filter_by(producto_id=prod_id)
    if fecha_desde:
        try:
            query = query.filter(MovimientoStock.fecha >= datetime.strptime(fecha_desde, '%Y-%m-%d'))
        except ValueError:
            pass
    if fecha_hasta:
        try:
            hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(MovimientoStock.fecha < hasta)
        except ValueError:
            pass

    movs      = query.order_by(MovimientoStock.fecha.desc()).paginate(page=pagina, per_page=25)
    prods_all = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    return render_template('admin/movimientos.html',
                           movimientos=movs, productos=prods_all,
                           tipo_sel=tipo, prod_sel=prod_id,
                           fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
                           TIPOS=MovimientoStock.TIPOS)


@bp.route('/movimientos/nuevo', methods=['GET', 'POST'])
@admin_requerido
def nuevo_movimiento():
    form = FormMovimiento()
    form.producto_id.choices = [
        (p.id, f'{p.nombre} (Stock: {p.stock})')
        for p in Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    ]
    form.proveedor_id.choices = [(0, '— Sin proveedor —')] + [
        (p.id, p.nombre) for p in Proveedor.query.filter_by(activo=True).all()
    ]
    if form.validate_on_submit():
        prod        = Producto.query.get_or_404(form.producto_id.data)
        tipo        = form.tipo.data
        cant        = form.cantidad.data
        stock_antes = prod.stock

        if tipo in ('salida', 'venta'):
            if cant > prod.stock:
                flash(f'Stock insuficiente. Stock actual: {prod.stock}', 'danger')
                return render_template('admin/movimiento_form.html', form=form)
            prod.stock -= cant
        elif tipo in ('entrada', 'devolucion'):
            prod.stock += cant
        elif tipo == 'ajuste':
            prod.stock = cant

        mov = MovimientoStock(
            producto_id=prod.id, tipo=tipo,
            cantidad=cant, stock_antes=stock_antes, stock_despues=prod.stock,
            motivo=form.motivo.data, referencia=form.referencia.data,
            proveedor_id=form.proveedor_id.data if form.proveedor_id.data else None,
            usuario_id=current_user.id,
        )
        db.session.add(mov)
        db.session.commit()
        flash('Movimiento registrado correctamente.', 'success')
        return redirect(url_for('admin.movimientos'))

    prod_id_pre = request.args.get('prod', 0, type=int)
    if prod_id_pre:
        form.producto_id.data = prod_id_pre
    return render_template('admin/movimiento_form.html', form=form)


# ═══════════════════════════════════════════════════════════════════════════════
# VENTAS FÍSICAS
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/ventas-fisicas')
@admin_requerido
def ventas_fisicas():
    pagina      = request.args.get('pagina', 1, type=int)
    fecha_desde = request.args.get('desde', '')
    fecha_hasta = request.args.get('hasta', '')
    metodo      = request.args.get('metodo', '')

    query = VentaFisica.query
    if not request.args.get('todas'):
        query = query.filter_by(anulada=False)
    if metodo:
        query = query.filter_by(metodo_pago=metodo)
    if fecha_desde:
        try:
            query = query.filter(VentaFisica.fecha >= datetime.strptime(fecha_desde, '%Y-%m-%d'))
        except ValueError:
            pass
    if fecha_hasta:
        try:
            hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(VentaFisica.fecha < hasta)
        except ValueError:
            pass

    ventas_pag = query.order_by(VentaFisica.fecha.desc()).paginate(page=pagina, per_page=20)
    return render_template('admin/ventas_fisicas.html', ventas=ventas_pag,
                           fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
                           metodo_sel=metodo, METODOS=VentaFisica.METODOS_PAGO)


@bp.route('/ventas-fisicas/nueva', methods=['GET', 'POST'])
@admin_requerido
def nueva_venta():
    form = FormVenta()
    productos_activos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    if request.method == 'POST':
        try:
            raw   = request.form.get('items_json', '[]')
            items = json.loads(raw)
        except Exception:
            flash('Error al procesar los productos. Inténtalo de nuevo.', 'danger')
            return render_template('admin/venta_form.html', form=form, productos=productos_activos)

        if not items:
            flash('Agrega al menos un producto a la venta.', 'warning')
            return render_template('admin/venta_form.html', form=form, productos=productos_activos)

        descuento = float(form.descuento.data or 0)
        subtotal  = sum(float(i['subtotal']) for i in items)
        total     = max(subtotal - descuento, 0)

        venta = VentaFisica(
            numero_venta   = VentaFisica.generar_numero(),
            cliente_nombre = form.cliente_nombre.data or None,
            cliente_doc    = form.cliente_doc.data or None,
            metodo_pago    = form.metodo_pago.data,
            subtotal       = subtotal,
            descuento      = descuento,
            total          = total,
            notas          = form.notas.data,
            usuario_id     = current_user.id,
        )
        db.session.add(venta)
        db.session.flush()

        from collections import defaultdict
        cantidades_por_prod = defaultdict(int)
        for item in items:
            cantidades_por_prod[int(item['producto_id'])] += int(item['cantidad'])
        for prod_id_chk, cant_total in cantidades_por_prod.items():
            prod_chk = Producto.query.get(prod_id_chk)
            if not prod_chk or prod_chk.stock < cant_total:
                db.session.rollback()
                nombre_chk = prod_chk.nombre if prod_chk else f'ID {prod_id_chk}'
                flash(f'Stock insuficiente para: {nombre_chk}.', 'danger')
                return render_template('admin/venta_form.html', form=form, productos=productos_activos)

        for item in items:
            prod = Producto.query.get(item['producto_id'])
            if not prod:
                db.session.rollback()
                flash(f'Producto no encontrado.', 'danger')
                return render_template('admin/venta_form.html', form=form, productos=productos_activos)
            detalle = DetalleVenta(
                venta_id    = venta.id,
                producto_id = prod.id,
                cantidad    = int(item['cantidad']),
                precio_unit = float(item['precio_unit']),
                talla       = item.get('talla') or None,
                color       = item.get('color') or None,
                subtotal    = float(item['subtotal']),
            )
            stock_antes = prod.stock
            prod.stock -= int(item['cantidad'])
            mov = MovimientoStock(
                producto_id  = prod.id, tipo='venta',
                cantidad     = int(item['cantidad']),
                stock_antes  = stock_antes, stock_despues=prod.stock,
                motivo       = f'Venta física {venta.numero_venta}',
                referencia   = venta.numero_venta,
                usuario_id   = current_user.id,
            )
            db.session.add(detalle)
            db.session.add(mov)

        db.session.commit()
        try:
            from utils.notificaciones import enviar_resumen_venta, enviar_alerta_stock
            enviar_resumen_venta(venta)
            bajos = [d.producto for d in venta.detalles.all() if d.producto.stock_bajo]
            if bajos:
                enviar_alerta_stock(bajos)
        except Exception:
            pass
        flash(f'Venta {venta.numero_venta} registrada correctamente.', 'success')
        return redirect(url_for('admin.ver_venta', id=venta.id))

    return render_template('admin/venta_form.html', form=form, productos=productos_activos)


@bp.route('/ventas-fisicas/<int:id>')
@admin_requerido
def ver_venta(id):
    venta = VentaFisica.query.get_or_404(id)
    return render_template('admin/venta_detalle.html', venta=venta)


@bp.route('/ventas-fisicas/<int:id>/anular', methods=['POST'])
@admin_requerido
def anular_venta(id):
    venta = VentaFisica.query.get_or_404(id)
    if venta.anulada:
        flash('Esta venta ya fue anulada.', 'warning')
        return redirect(url_for('admin.ver_venta', id=id))
    venta.anulada = True
    for det in venta.detalles.all():
        prod        = det.producto
        stock_antes = prod.stock
        prod.stock += det.cantidad
        mov = MovimientoStock(
            producto_id=prod.id, tipo='devolucion',
            cantidad=det.cantidad, stock_antes=stock_antes, stock_despues=prod.stock,
            motivo=f'Anulación venta {venta.numero_venta}',
            referencia=venta.numero_venta, usuario_id=current_user.id,
        )
        db.session.add(mov)
    db.session.commit()
    flash(f'Venta {venta.numero_venta} anulada y stock revertido.', 'info')
    return redirect(url_for('admin.ventas_fisicas'))


# ═══════════════════════════════════════════════════════════════════════════════
# PEDIDOS ONLINE
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/pedidos')
@admin_requerido
def pedidos():
    pagina = request.args.get('pagina', 1, type=int)
    estado = request.args.get('estado', '')
    query  = Pedido.query
    if estado:
        query = query.filter_by(estado=estado)
    peds = query.order_by(Pedido.fecha_creacion.desc()).paginate(page=pagina, per_page=20)
    return render_template('admin/pedidos.html', pedidos=peds,
                           estado_sel=estado, ESTADOS=Pedido.ESTADOS)


@bp.route('/pedidos/<int:id>')
@admin_requerido
def ver_pedido(id):
    pedido = Pedido.query.get_or_404(id)
    return render_template('admin/pedido_detalle.html', pedido=pedido)


@bp.route('/pedidos/<int:id>/eliminar', methods=['POST'])
@admin_requerido
def eliminar_pedido(id):
    pedido = Pedido.query.get_or_404(id)
    numero = pedido.numero_pedido
    # Borrar detalles manualmente primero (por si MySQL no tiene ON DELETE CASCADE aún)
    from models.pedido import DetallePedido
    DetallePedido.query.filter_by(pedido_id=id).delete()
    db.session.delete(pedido)
    db.session.commit()
    flash(f'Pedido {numero} eliminado correctamente.', 'success')
    return redirect(url_for('admin.pedidos'))


@bp.route('/pedidos/<int:id>/estado', methods=['POST'])
@admin_requerido
def cambiar_estado(id):
    pedido = Pedido.query.get_or_404(id)
    nuevo_estado = request.form.get('estado')
    if nuevo_estado in Pedido.ESTADOS:
        pedido.estado = nuevo_estado
        if nuevo_estado in ('entregado', 'confirmado') and pedido.metodo_pago in ('efectivo', 'recojo'):
            pedido.estado_pago = 'pagado'
        db.session.commit()
        flash(f'Estado actualizado a: {pedido.estado_label}', 'success')
    return redirect(url_for('admin.ver_pedido', id=id))


# ═══════════════════════════════════════════════════════════════════════════════
# CLIENTES
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/clientes')
@admin_requerido
def clientes():
    pagina   = request.args.get('pagina', 1, type=int)
    busqueda = request.args.get('q', '')
    query    = Usuario.query.filter_by(es_admin=False)
    if busqueda:
        query = query.filter(
            (Usuario.nombre.ilike(f'%{busqueda}%')) |
            (Usuario.email.ilike(f'%{busqueda}%'))
        )
    clientes_pag = query.order_by(Usuario.fecha_registro.desc()).paginate(page=pagina, per_page=20)
    return render_template('admin/clientes.html', clientes=clientes_pag, busqueda=busqueda)


@bp.route('/clientes/<int:id>')
@admin_requerido
def ver_cliente(id):
    cliente = Usuario.query.get_or_404(id)
    peds    = Pedido.query.filter_by(usuario_id=id).order_by(Pedido.fecha_creacion.desc()).all()
    return render_template('admin/cliente_detalle.html', cliente=cliente, pedidos=peds)


@bp.route('/clientes/<int:id>/toggle', methods=['POST'])
@admin_requerido
def toggle_cliente(id):
    cliente = Usuario.query.get_or_404(id)
    cliente.activo = not cliente.activo
    db.session.commit()
    estado = 'activado' if cliente.activo else 'desactivado'
    flash(f'Cliente {estado}.', 'info')
    return redirect(url_for('admin.clientes'))


@bp.route('/clientes/<int:id>/eliminar', methods=['POST'])
@admin_requerido
def eliminar_cliente(id):
    cliente = Usuario.query.get_or_404(id)

    if cliente.es_admin:
        flash('No puedes eliminar un administrador.', 'danger')
        return redirect(url_for('admin.clientes'))

    if cliente.id == current_user.id:
        flash('No puedes eliminar tu propia cuenta.', 'danger')
        return redirect(url_for('admin.clientes'))

    nombre = cliente.nombre_completo
    # Borrar en cascada manualmente (MySQL puede no tener ON DELETE CASCADE aún)
    from models.pedido import DetallePedido
    pedido_ids = [p.id for p in cliente.pedidos.all()]
    if pedido_ids:
        DetallePedido.query.filter(DetallePedido.pedido_id.in_(pedido_ids)).delete(synchronize_session=False)
        Pedido.query.filter(Pedido.usuario_id == id).delete(synchronize_session=False)
    db.session.delete(cliente)
    db.session.commit()
    flash(f'Cliente "{nombre}" eliminado correctamente.', 'success')
    return redirect(url_for('admin.clientes'))


# ═══════════════════════════════════════════════════════════════════════════════
# CATEGORÍAS
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/categorias')
@admin_requerido
def categorias():
    cats = Categoria.query.order_by(Categoria.nombre).all()
    return render_template('admin/categorias.html', categorias=cats)


@bp.route('/categorias/nueva', methods=['GET', 'POST'])
@admin_requerido
def nueva_categoria():
    form = FormCategoria()
    if form.validate_on_submit():
        slug = form.slug.data or _slugify(form.nombre.data)
        cat  = Categoria(nombre=form.nombre.data, slug=slug,
                         descripcion=form.descripcion.data,
                         icono=form.icono.data or 'box-seam',
                         activo=form.activo.data)
        db.session.add(cat)
        db.session.commit()
        flash('Categoría creada.', 'success')
        return redirect(url_for('admin.categorias'))
    return render_template('admin/categoria_form.html', form=form, titulo='Nueva categoría')


@bp.route('/categorias/<int:id>/editar', methods=['GET', 'POST'])
@admin_requerido
def editar_categoria(id):
    cat  = Categoria.query.get_or_404(id)
    form = FormCategoria(obj=cat)
    if form.validate_on_submit():
        cat.nombre      = form.nombre.data
        cat.slug        = form.slug.data or _slugify(form.nombre.data)
        cat.descripcion = form.descripcion.data
        cat.icono       = form.icono.data or 'box-seam'
        cat.activo      = form.activo.data
        db.session.commit()
        flash('Categoría actualizada.', 'success')
        return redirect(url_for('admin.categorias'))
    return render_template('admin/categoria_form.html', form=form,
                           titulo='Editar categoría', categoria=cat)


# ═══════════════════════════════════════════════════════════════════════════════
# PROVEEDORES
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/proveedores')
@admin_requerido
def proveedores():
    busqueda = request.args.get('q', '')
    query    = Proveedor.query
    if busqueda:
        query = query.filter(
            (Proveedor.nombre.ilike(f'%{busqueda}%')) |
            (Proveedor.ruc.ilike(f'%{busqueda}%'))
        )
    provs = query.order_by(Proveedor.nombre).all()
    return render_template('admin/proveedores.html', proveedores=provs, busqueda=busqueda)


@bp.route('/proveedores/nuevo', methods=['GET', 'POST'])
@admin_requerido
def nuevo_proveedor():
    form = FormProveedor()
    if form.validate_on_submit():
        prov = Proveedor(
            nombre=form.nombre.data, ruc=form.ruc.data or None,
            contacto=form.contacto.data, telefono=form.telefono.data,
            email=form.email.data, direccion=form.direccion.data,
            notas=form.notas.data, activo=form.activo.data,
        )
        db.session.add(prov)
        db.session.commit()
        flash('Proveedor registrado.', 'success')
        return redirect(url_for('admin.proveedores'))
    return render_template('admin/proveedor_form.html', form=form, titulo='Nuevo proveedor')


@bp.route('/proveedores/<int:id>/editar', methods=['GET', 'POST'])
@admin_requerido
def editar_proveedor(id):
    prov = Proveedor.query.get_or_404(id)
    form = FormProveedor(obj=prov)
    if form.validate_on_submit():
        prov.nombre    = form.nombre.data
        prov.ruc       = form.ruc.data or None
        prov.contacto  = form.contacto.data
        prov.telefono  = form.telefono.data
        prov.email     = form.email.data
        prov.direccion = form.direccion.data
        prov.notas     = form.notas.data
        prov.activo    = form.activo.data
        db.session.commit()
        flash('Proveedor actualizado.', 'success')
        return redirect(url_for('admin.proveedores'))
    return render_template('admin/proveedor_form.html', form=form,
                           titulo='Editar proveedor', proveedor=prov)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTES
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/reportes')
@admin_requerido
def reportes():
    hoy   = datetime.utcnow().date()
    mes_i = datetime(hoy.year, hoy.month, 1)

    # ── Ventas por mes: suma VentaFisica + Pedidos pagados ────────────────────
    ventas_por_mes = []
    for i in range(5, -1, -1):
        if hoy.month - i <= 0:
            m = hoy.month - i + 12
            y = hoy.year - 1
        else:
            m = hoy.month - i
            y = hoy.year
        desde = datetime(y, m, 1)
        hasta = datetime(y + 1, 1, 1) if m == 12 else datetime(y, m + 1, 1)

        total_fisico = db.session.query(func.sum(VentaFisica.total))\
            .filter(VentaFisica.fecha >= desde, VentaFisica.fecha < hasta,
                    VentaFisica.anulada == False).scalar() or 0

        total_online = db.session.query(func.sum(Pedido.total))\
            .filter(Pedido.fecha_creacion >= desde, Pedido.fecha_creacion < hasta,
                    Pedido.estado_pago == 'pagado',
                    Pedido.estado != 'cancelado').scalar() or 0

        ventas_por_mes.append({
            'mes':   desde.strftime('%b %Y'),
            'total': float(total_fisico) + float(total_online),
        })

    # ── Métodos de pago del mes: VentaFisica + Pedidos ───────────────────────
    # Ventas físicas del mes
    metodos_fisico = db.session.query(
        VentaFisica.metodo_pago,
        func.sum(VentaFisica.total).label('total'),
        func.count(VentaFisica.id).label('cantidad')
    ).filter(VentaFisica.fecha >= mes_i, VentaFisica.anulada == False)\
     .group_by(VentaFisica.metodo_pago).all()

    # Pedidos online pagados del mes
    metodos_online = db.session.query(
        Pedido.metodo_pago,
        func.sum(Pedido.total).label('total'),
        func.count(Pedido.id).label('cantidad')
    ).filter(Pedido.fecha_creacion >= mes_i,
             Pedido.estado_pago == 'pagado',
             Pedido.estado != 'cancelado')\
     .group_by(Pedido.metodo_pago).all()

    # Combinar los dos en un dict para agrupar por método
    metodos_dict = {}
    for row in list(metodos_fisico) + list(metodos_online):
        key = row.metodo_pago
        if key not in metodos_dict:
            metodos_dict[key] = {'metodo_pago': key, 'total': 0.0, 'cantidad': 0}
        metodos_dict[key]['total']    += float(row.total or 0)
        metodos_dict[key]['cantidad'] += int(row.cantidad or 0)

    # Convertir a lista de objetos tipo namedtuple para el template
    from collections import namedtuple
    MetodoPago = namedtuple('MetodoPago', ['metodo_pago', 'total', 'cantidad'])
    ventas_metodo = [MetodoPago(**v) for v in metodos_dict.values()]

    # ── Top 10 productos del mes: VentaFisica + Pedidos ──────────────────────
    # Desde ventas físicas
    top_fisico = db.session.query(
        Producto.id,
        Producto.nombre,
        Producto.sku,
        func.sum(DetalleVenta.cantidad).label('unidades'),
        func.sum(DetalleVenta.subtotal).label('monto')
    ).join(DetalleVenta, DetalleVenta.producto_id == Producto.id)\
     .join(VentaFisica, VentaFisica.id == DetalleVenta.venta_id)\
     .filter(VentaFisica.fecha >= mes_i, VentaFisica.anulada == False)\
     .group_by(Producto.id, Producto.nombre, Producto.sku).all()

    # Desde pedidos online
    top_online = db.session.query(
        Producto.id,
        Producto.nombre,
        Producto.sku,
        func.sum(DetallePedido.cantidad).label('unidades'),
        func.sum(DetallePedido.subtotal).label('monto')
    ).join(DetallePedido, DetallePedido.producto_id == Producto.id)\
     .join(Pedido, Pedido.id == DetallePedido.pedido_id)\
     .filter(Pedido.fecha_creacion >= mes_i,
             Pedido.estado_pago == 'pagado',
             Pedido.estado != 'cancelado')\
     .group_by(Producto.id, Producto.nombre, Producto.sku).all()

    # Combinar por producto_id
    top_dict = {}
    for row in list(top_fisico) + list(top_online):
        pid = row.id
        if pid not in top_dict:
            top_dict[pid] = {'nombre': row.nombre, 'sku': row.sku,
                             'unidades': 0, 'monto': 0.0}
        top_dict[pid]['unidades'] += int(row.unidades or 0)
        top_dict[pid]['monto']    += float(row.monto or 0)

    TopProducto = namedtuple('TopProducto', ['nombre', 'sku', 'unidades', 'monto'])
    top_productos = sorted(
        [TopProducto(**v) for v in top_dict.values()],
        key=lambda x: x.monto, reverse=True
    )[:10]

    # ── Inventario por categoría (sin cambios) ────────────────────────────────
    inventario_cat = db.session.query(
        Categoria.nombre,
        func.count(Producto.id).label('productos'),
        func.sum(Producto.stock).label('unidades'),
        func.sum(Producto.precio * Producto.stock).label('valor')
    ).join(Producto, Producto.categoria_id == Categoria.id)\
     .filter(Producto.activo == True)\
     .group_by(Categoria.id, Categoria.nombre).all()

    sin_stock  = Producto.query.filter_by(activo=True).filter(Producto.stock <= 0).count()
    stock_bajo = Producto.query.filter(Producto.stock > 0,
                                       Producto.stock <= Producto.stock_minimo,
                                       Producto.activo == True).count()

    return render_template('admin/reportes.html',
        ventas_por_mes=ventas_por_mes, ventas_metodo=ventas_metodo,
        top_productos=top_productos, inventario_cat=inventario_cat,
        sin_stock=sin_stock, stock_bajo=stock_bajo,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# NOTIFICACIONES
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/enviar-resumen-diario', methods=['POST'])
@admin_requerido
def enviar_resumen_diario_manual():
    from utils.notificaciones import enviar_resumen_diario
    hoy = datetime.utcnow().date()
    ventas_hoy = VentaFisica.query\
        .filter(func.date(VentaFisica.fecha) == hoy, VentaFisica.anulada == False).all()
    total_hoy = sum(float(v.total) for v in ventas_hoy)
    productos_bajos = Producto.query.filter(
        Producto.stock <= Producto.stock_minimo, Producto.activo == True
    ).all()
    try:
        enviar_resumen_diario(ventas_hoy, total_hoy, productos_bajos)
        flash('Resumen del día enviado al correo.', 'success')
    except Exception as e:
        flash(f'Error enviando el resumen: {e}', 'danger')
    return redirect(url_for('admin.dashboard'))


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORTAR EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/exportar/productos')
@admin_requerido
def exportar_productos():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    prods = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos'
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill(fill_type='solid', fgColor='A53694')
    ha = Alignment(horizontal='center')
    encabezados = ['#','Nombre','SKU','Categoría','Proveedor',
                   'P. Compra','P. Venta','Stock','Stock Mín','Margen %','Valor Inv.','Estado']
    for col, t in enumerate(encabezados, 1):
        c = ws.cell(row=1, column=col, value=t)
        c.font = hf; c.fill = hfill; c.alignment = ha
    for row, p in enumerate(prods, 2):
        ws.cell(row=row, column=1,  value=row-1)
        ws.cell(row=row, column=2,  value=p.nombre)
        ws.cell(row=row, column=3,  value=p.sku or '—')
        ws.cell(row=row, column=4,  value=p.categoria_rel.nombre if p.categoria_rel else '—')
        ws.cell(row=row, column=5,  value=p.proveedor.nombre if p.proveedor else '—')
        ws.cell(row=row, column=6,  value=float(p.precio_compra or 0))
        ws.cell(row=row, column=7,  value=float(p.precio))
        ws.cell(row=row, column=8,  value=p.stock)
        ws.cell(row=row, column=9,  value=p.stock_minimo)
        ws.cell(row=row, column=10, value=p.margen or 0)
        ws.cell(row=row, column=11, value=p.valor_inventario)
        ws.cell(row=row, column=12, value='Activo')
        if p.sin_stock:
            ws.cell(row=row, column=8).font = Font(bold=True, color='CC0000')
        elif p.stock_bajo:
            ws.cell(row=row, column=8).font = Font(bold=True, color='FF8800')
    anchos = [4,35,15,18,20,14,14,10,12,10,16,10]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    fecha = datetime.utcnow().strftime('%Y%m%d')
    return Response(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=productos_{fecha}.xlsx'})


@bp.route('/exportar/ventas')
@admin_requerido
def exportar_ventas():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    ventas = VentaFisica.query.filter_by(anulada=False).order_by(VentaFisica.fecha.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Ventas'
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill(fill_type='solid', fgColor='A53694')
    ha = Alignment(horizontal='center')
    encabezados = ['#','N° Venta','Fecha','Cliente','Doc.','Método Pago','Subtotal','Descuento','Total','Notas']
    for col, t in enumerate(encabezados, 1):
        c = ws.cell(row=1, column=col, value=t)
        c.font = hf; c.fill = hfill; c.alignment = ha
    total_general = 0
    for row, v in enumerate(ventas, 2):
        ws.cell(row=row, column=1,  value=row-1)
        ws.cell(row=row, column=2,  value=v.numero_venta)
        ws.cell(row=row, column=3,  value=v.fecha.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=4,  value=v.cliente_nombre or '—')
        ws.cell(row=row, column=5,  value=v.cliente_doc or '—')
        ws.cell(row=row, column=6,  value=v.metodo_pago_label)
        ws.cell(row=row, column=7,  value=float(v.subtotal))
        ws.cell(row=row, column=8,  value=float(v.descuento))
        ws.cell(row=row, column=9,  value=float(v.total))
        ws.cell(row=row, column=10, value=v.notas or '')
        total_general += float(v.total)
    fila_total = len(ventas) + 2
    ws.cell(row=fila_total, column=8, value='TOTAL:').font = Font(bold=True)
    ws.cell(row=fila_total, column=9, value=total_general).font = Font(bold=True, color='A53694')
    anchos = [4,22,18,25,12,16,12,12,12,30]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    fecha = datetime.utcnow().strftime('%Y%m%d')
    return Response(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=ventas_{fecha}.xlsx'})


@bp.route('/exportar/movimientos')
@admin_requerido
def exportar_movimientos():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    movs = MovimientoStock.query.order_by(MovimientoStock.fecha.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Movimientos'
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill(fill_type='solid', fgColor='A53694')
    ha = Alignment(horizontal='center')
    encabezados = ['#','Fecha','Tipo','Producto','Cantidad','Stock Ant.','Stock Nuevo','Motivo','Referencia','Proveedor','Usuario']
    for col, t in enumerate(encabezados, 1):
        c = ws.cell(row=1, column=col, value=t)
        c.font = hf; c.fill = hfill; c.alignment = ha
    for row, m in enumerate(movs, 2):
        ws.cell(row=row, column=1,  value=row-1)
        ws.cell(row=row, column=2,  value=m.fecha.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=3,  value=m.tipo_label)
        ws.cell(row=row, column=4,  value=m.producto.nombre if m.producto else '—')
        ws.cell(row=row, column=5,  value=m.cantidad)
        ws.cell(row=row, column=6,  value=m.stock_antes)
        ws.cell(row=row, column=7,  value=m.stock_despues)
        ws.cell(row=row, column=8,  value=m.motivo or '—')
        ws.cell(row=row, column=9,  value=m.referencia or '—')
        ws.cell(row=row, column=10, value=m.proveedor.nombre if m.proveedor else '—')
        ws.cell(row=row, column=11, value=m.usuario.nombre if m.usuario else '—')
    anchos = [4,18,12,30,10,10,12,35,20,25,20]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    fecha = datetime.utcnow().strftime('%Y%m%d')
    return Response(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=movimientos_{fecha}.xlsx'})


# ═══════════════════════════════════════════════════════════════════════════════
# API AJAX
# ═══════════════════════════════════════════════════════════════════════════════
@bp.route('/api/producto/<int:id>')
@admin_requerido
def api_producto(id):
    prod = Producto.query.get_or_404(id)
    return jsonify({
        'id': prod.id, 'nombre': prod.nombre,
        'precio_venta': float(prod.precio),
        'stock': prod.stock, 'sku': prod.sku or '',
        'tallas': prod.tallas_lista, 'colores': prod.colores_lista,
        'unidad': prod.unidad,
    })


@bp.route('/api/buscar-productos')
@admin_requerido
def api_buscar_productos():
    q     = request.args.get('q', '')
    prods = Producto.query.filter(
        Producto.activo==True,
        (Producto.nombre.ilike(f'%{q}%')) | (Producto.sku.ilike(f'%{q}%'))
    ).limit(10).all()
    return jsonify([{
        'id': p.id, 'nombre': p.nombre,
        'precio_venta': float(p.precio),
        'stock': p.stock, 'sku': p.sku or '',
    } for p in prods])
